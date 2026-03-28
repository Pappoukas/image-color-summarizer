#!/usr/bin/env python3
"""
batch_from_urls.py
==================
Reads image URLs from a .txt file (local or on GitHub),
analyses the colors of each image, and saves all results
plus the original URL into a formatted Excel file.

Usage:
    python batch_from_urls.py urls.txt
    python batch_from_urls.py https://raw.githubusercontent.com/USER/REPO/main/urls.txt
    python batch_from_urls.py urls.txt -k 6 -o results/
    python batch_from_urls.py urls.txt -k 8 --no-png

URL file format (one URL per line; lines starting with # are ignored):
    https://example.com/photo1.jpg
    # this is a comment
    https://example.com/photo2.png
"""

import argparse
import io
import sys
import tempfile
import time
import traceback
import urllib.request
from pathlib import Path

import numpy as np
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from PIL import Image

# Import the colour engine from the companion script
sys.path.insert(0, str(Path(__file__).parent))
from color_summarizer import (
    analyze,
    cluster_pixels,
    image_statistics,
    rgb_to_hsv,
    rgb_to_lab,
    lab_to_lch,
    color_name,
)


# ─────────────────────────────────────────────────────────────────────────────
# URL helpers
# ─────────────────────────────────────────────────────────────────────────────

def _raw_github(url: str) -> str:
    """Convert a regular GitHub blob URL to raw.githubusercontent.com."""
    return url.replace(
        "github.com", "raw.githubusercontent.com"
    ).replace("/blob/", "/")


def load_url_list(source: str) -> list[str]:
    """
    Load URLs from a local file path OR a remote URL.
    Lines starting with # (or empty) are skipped.
    """
    if source.startswith("http://") or source.startswith("https://"):
        raw = _raw_github(source)
        print(f"[→] Fetching URL list from: {raw}")
        r = requests.get(raw, timeout=30)
        r.raise_for_status()
        lines = r.text.splitlines()
    else:
        with open(source, encoding="utf-8") as fh:
            lines = fh.read().splitlines()

    urls = [ln.strip() for ln in lines
            if ln.strip() and not ln.strip().startswith("#")]
    print(f"[→] Found {len(urls)} URL(s)")
    return urls


def download_image(url: str) -> Image.Image:
    """Download an image from a URL and return a PIL Image (RGB)."""
    raw = _raw_github(url)
    r = requests.get(raw, timeout=30)
    r.raise_for_status()
    img = Image.open(io.BytesIO(r.content)).convert("RGB")
    return img


# ─────────────────────────────────────────────────────────────────────────────
# Excel helpers
# ─────────────────────────────────────────────────────────────────────────────

# ── Colour constants ──────────────────────────────────────────────────────────
DARK_BG   = "1C2833"   # header rows
MID_BG    = "2E4057"   # sub-header rows
ACCENT    = "1ABC9C"   # teal accent line / cluster header
ROW_ODD   = "F4F6F7"
ROW_EVEN  = "FDFEFE"
LINK_FONT = "2980B9"

THIN  = Side(style="thin",   color="CCCCCC")
THICK = Side(style="medium", color="999999")

def _border(all_thin=False) -> Border:
    s = THIN
    return Border(left=s, right=s, top=s, bottom=s)

def _header_font(size=11, bold=True, color="FFFFFF"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def _body_font(size=9, bold=False, color="2C3E50"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _center(wrap=False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=False) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def _hex_to_rgb_tuple(hex_color: str) -> tuple:
    h = hex_color.lstrip("#")
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))


def _readable_on(hex_color: str) -> str:
    r, g, b = _hex_to_rgb_tuple(hex_color)
    return "000000" if (0.299*r + 0.587*g + 0.114*b) > 140 else "FFFFFF"


def _set_col_widths(ws, widths: dict) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _write_cell(ws, row, col, value,
                font=None, fill=None, alignment=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:      c.font      = font
    if fill:      c.fill      = fill
    if alignment: c.alignment = alignment
    if border:    c.border    = border
    return c


# ─────────────────────────────────────────────────────────────────────────────
# Build the workbook
# ─────────────────────────────────────────────────────────────────────────────

def build_excel(results: list[dict], output_path: Path) -> None:
    """
    Build and save the Excel workbook from the list of result dicts.

    Each dict has keys:
        url, filename, width, height, status,
        stats (or None), clusters (list or []),
        error (str or None)
    """
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _build_summary_sheet(ws_sum, results)

    # ── Sheet 2: Clusters ─────────────────────────────────────────────────────
    ws_cl = wb.create_sheet("Clusters")
    _build_clusters_sheet(ws_cl, results)

    # ── Sheet 3: Statistics ───────────────────────────────────────────────────
    ws_st = wb.create_sheet("Statistics")
    _build_stats_sheet(ws_st, results)

    wb.save(output_path)
    print(f"[✓] Excel saved: {output_path}")


# ── Summary sheet ─────────────────────────────────────────────────────────────
#
# Layout (3 header rows + data):
#
# Row 1 : Title  (merged)
# Row 2 : Group labels  — "Image Info" | "Mean Values" | "Color 1" … "Color 5" | "Status"
# Row 3 : Column headers
# Row 4+: Data
#
# Fixed columns  (A–L = 1–12):
#   1  #          2  URL        3  Filename   4  W×H
#   5  R mean     6  G mean     7  B mean
#   8  H° mean    9  S% mean   10  V% mean
#  11  L mean    12  C mean
#
# Color columns, 3 per cluster × 5 clusters = 15 cols  (M–AA = 13–27):
#   13  C1 swatch   14  C1 %   15  C1 name
#   16  C2 swatch   17  C2 %   18  C2 name
#   … etc.
#
# Status column : 28 (AB)

_N_TOP        = 5          # number of top colors to show
_COL_FIXED    = 12         # last fixed column index
_COLS_PER_CL  = 3          # swatch + % + name
_COL_STATUS   = _COL_FIXED + _N_TOP * _COLS_PER_CL + 1   # = 28

# Pastel accent colours for the 5 cluster group headers
_CL_HEADER_FILLS = ["1A5276", "1F618D", "2471A3", "2980B9", "5499C9"]


def _build_summary_sheet(ws, results):
    total_cols = _COL_STATUS

    # ── Row 1: Title ───────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=total_cols)
    c = ws.cell(row=1, column=1,
                value="IMAGE COLOR SUMMARIZER  —  Batch Results")
    c.font      = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    c.fill      = _fill(DARK_BG)
    c.alignment = _center()
    ws.row_dimensions[1].height = 30

    # ── Row 2: Group labels ────────────────────────────────────────────────
    def _group(c1, c2, label, bg):
        ws.merge_cells(start_row=2, start_column=c1,
                       end_row=2,   end_column=c2)
        cell = ws.cell(row=2, column=c1, value=label)
        cell.font      = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        cell.fill      = _fill(bg)
        cell.alignment = _center()
        cell.border    = _border()
        for col in range(c1 + 1, c2 + 1):
            ws.cell(row=2, column=col).fill   = _fill(bg)
            ws.cell(row=2, column=col).border = _border()

    _group(1, 4,   "Image Info",    DARK_BG)
    _group(5, 12,  "Mean Values",   MID_BG)
    for i in range(_N_TOP):
        c1 = _COL_FIXED + i * _COLS_PER_CL + 1
        c2 = c1 + _COLS_PER_CL - 1
        _group(c1, c2, f"Color {i+1}", _CL_HEADER_FILLS[i])
    ws.cell(row=2, column=_COL_STATUS,
            value="").fill = _fill(DARK_BG)
    ws.row_dimensions[2].height = 18

    # ── Row 3: Column headers ──────────────────────────────────────────────
    col_headers = (
        ["#", "URL", "Filename", "W×H",
         "R mean", "G mean", "B mean",
         "H° mean", "S% mean", "V% mean",
         "L mean", "C mean"]
        + ["HEX", "%", "Name"] * _N_TOP
        + ["Status"]
    )
    for ci, h in enumerate(col_headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font      = _header_font(size=9)
        c.fill      = _fill(MID_BG)
        c.alignment = _center(wrap=True)
        c.border    = _border()
    ws.row_dimensions[3].height = 22

    ws.freeze_panes = "A4"

    # ── Data rows ──────────────────────────────────────────────────────────
    for ri, res in enumerate(results, 1):
        row = ri + 3
        bg  = ROW_ODD if ri % 2 else ROW_EVEN
        ok  = res["status"] == "ok"

        def wc(col, val, bold=False, align="center",
               color="2C3E50", wrap=False, cell_bg=None):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = Font(name="Arial", size=9, bold=bold, color=color)
            c.fill      = _fill(cell_bg if cell_bg else bg)
            c.alignment = Alignment(horizontal=align, vertical="center",
                                    wrap_text=wrap)
            c.border    = _border()

        wc(1, ri, bold=True)

        c = ws.cell(row=row, column=2, value=res["url"])
        c.hyperlink  = res["url"]
        c.font       = Font(name="Arial", size=9, color=LINK_FONT,
                            underline="single")
        c.fill       = _fill(bg)
        c.alignment  = _left(wrap=True)
        c.border     = _border()

        wc(3,  res["filename"], align="left")
        wc(4,  f"{res['width']}×{res['height']}" if ok else "—")

        if ok and res["stats"]:
            st = res["stats"]
            wc(5,  round(st["RGB"]["R"]["mean"], 1))
            wc(6,  round(st["RGB"]["G"]["mean"], 1))
            wc(7,  round(st["RGB"]["B"]["mean"], 1))
            wc(8,  round(st["HSV"]["H"]["mean"], 1))
            wc(9,  round(st["HSV"]["S"]["mean"], 1))
            wc(10, round(st["HSV"]["V"]["mean"], 1))
            wc(11, round(st["LCH"]["L"]["mean"], 1))
            wc(12, round(st["LCH"]["C"]["mean"], 1))

        # ── 5 color columns ───────────────────────────────────────────────
        clusters = res.get("clusters", []) if ok else []
        for rank in range(_N_TOP):
            base = _COL_FIXED + rank * _COLS_PER_CL + 1   # swatch col
            if rank < len(clusters):
                cl      = clusters[rank]
                hex_str = cl["hex"].lstrip("#")
                txt_col = _readable_on(hex_str)

                # Swatch cell: coloured background with hex text
                c = ws.cell(row=row, column=base, value=cl["hex"])
                c.font      = Font(name="Arial", size=9, bold=True, color=txt_col)
                c.fill      = _fill(hex_str)
                c.alignment = _center()
                c.border    = _border()

                # Percentage
                wc(base + 1, cl["percentage"])

                # Name
                wc(base + 2, cl["name"], align="left")
            else:
                for offset in range(_COLS_PER_CL):
                    wc(base + offset, "—")

        # Status
        status_color = "27AE60" if ok else "E74C3C"
        err_text = res.get("error", "ERR") or "ERR"
        c = ws.cell(row=row, column=_COL_STATUS,
                    value="✓ OK" if ok else f"✗ {err_text[:35]}")
        c.font      = Font(name="Arial", size=9, bold=True, color=status_color)
        c.fill      = _fill(bg)
        c.alignment = _center(wrap=True)
        c.border    = _border()

        ws.row_dimensions[row].height = 32

    # ── Column widths ──────────────────────────────────────────────────────
    widths = {
        "A": 5,  "B": 44, "C": 20, "D": 10,
        "E": 8,  "F": 8,  "G": 8,
        "H": 8,  "I": 8,  "J": 8,
        "K": 8,  "L": 8,
    }
    # Color columns: swatch=10, %=7, name=18 — repeat for each cluster
    from openpyxl.utils import get_column_letter
    for rank in range(_N_TOP):
        base = _COL_FIXED + rank * _COLS_PER_CL + 1
        widths[get_column_letter(base)]     = 10   # swatch/hex
        widths[get_column_letter(base + 1)] = 7    # %
        widths[get_column_letter(base + 2)] = 18   # name
    widths[get_column_letter(_COL_STATUS)] = 22
    _set_col_widths(ws, widths)


# ── Clusters sheet ────────────────────────────────────────────────────────────

def _build_clusters_sheet(ws, results):
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:R1")
    c = ws["A1"]
    c.value     = "COLOR CLUSTERS  —  per image"
    c.font      = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    c.fill      = _fill(DARK_BG)
    c.alignment = _center()
    ws.row_dimensions[1].height = 26

    headers = [
        "#", "URL", "Filename",
        "Cluster", "Pixels", "%",
        "HEX", "Swatch",
        "R", "G", "B",
        "H°", "S%", "V%",
        "L*", "a*", "b*",
        "Name",
    ]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font      = _header_font(size=9)
        c.fill      = _fill(MID_BG)
        c.alignment = _center(wrap=True)
        c.border    = _border()
    ws.row_dimensions[2].height = 22

    row = 3
    for img_i, res in enumerate(results, 1):
        if res["status"] != "ok" or not res["clusters"]:
            continue
        for cl in res["clusters"]:
            bg  = ROW_ODD if img_i % 2 else ROW_EVEN
            hex_str = cl["hex"].lstrip("#")

            def wc(col, val, bold=False, color="2C3E50"):
                c = ws.cell(row=row, column=col, value=val)
                c.font      = Font(name="Arial", size=9, bold=bold, color=color)
                c.fill      = _fill(bg)
                c.alignment = _center()
                c.border    = _border()

            wc(1,  img_i, bold=True)
            c2 = ws.cell(row=row, column=2, value=res["url"])
            c2.hyperlink  = res["url"]
            c2.font       = Font(name="Arial", size=9, color=LINK_FONT,
                                 underline="single")
            c2.fill       = _fill(bg)
            c2.alignment  = _left(wrap=True)
            c2.border     = _border()

            c3 = ws.cell(row=row, column=3, value=res["filename"])
            c3.font       = Font(name="Arial", size=9, color="2C3E50")
            c3.fill       = _fill(bg)
            c3.alignment  = _left()
            c3.border     = _border()

            wc(4,  cl["index"] + 1)
            wc(5,  cl["pixels"])
            wc(6,  cl["percentage"])

            # HEX cell with swatch colour
            txt_col = _readable_on(hex_str)
            c7 = ws.cell(row=row, column=7, value=cl["hex"])
            c7.font      = Font(name="Arial", size=9, bold=True, color=txt_col)
            c7.fill      = _fill(hex_str)
            c7.alignment = _center()
            c7.border    = _border()

            # Swatch (wider coloured cell)
            c8 = ws.cell(row=row, column=8, value="")
            c8.fill   = _fill(hex_str)
            c8.border = _border()

            r_, g_, b_ = cl["rgb"]
            h_, s_, v_ = cl["hsv"]
            la, aa, ba = cl["lab"]
            wc(9,  r_);  wc(10, g_);  wc(11, b_)
            wc(12, int(h_)); wc(13, int(s_)); wc(14, int(v_))
            wc(15, round(la, 1)); wc(16, round(aa, 1)); wc(17, round(ba, 1))

            c18 = ws.cell(row=row, column=18, value=cl["name"])
            c18.font      = Font(name="Arial", size=9, color="2C3E50",
                                 italic=True)
            c18.fill      = _fill(bg)
            c18.alignment = _left()
            c18.border    = _border()

            ws.row_dimensions[row].height = 18
            row += 1

    _set_col_widths(ws, {
        "A": 5,  "B": 38, "C": 20,
        "D": 8,  "E": 9,  "F": 7,
        "G": 10, "H": 6,
        "I": 5,  "J": 5,  "K": 5,
        "L": 6,  "M": 6,  "N": 6,
        "O": 7,  "P": 7,  "Q": 7,
        "R": 22,
    })


# ── Statistics sheet ──────────────────────────────────────────────────────────

def _build_stats_sheet(ws, results):
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value     = "COLOR SPACE STATISTICS  —  per image / per channel"
    c.font      = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    c.fill      = _fill(DARK_BG)
    c.alignment = _center()
    ws.row_dimensions[1].height = 26

    # Sub-header: spaces
    space_cols = {
        "RGB":  (4, 6),
        "HSV":  (7, 9),
        "Lab":  (10, 12),
        "LCH":  (13, 15),
    }
    for space, (c1, c2) in space_cols.items():
        ws.merge_cells(
            start_row=2, start_column=c1,
            end_row=2,   end_column=c2
        )
        c = ws.cell(row=2, column=c1, value=space)
        c.font      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill      = _fill(ACCENT)
        c.alignment = _center()
    for col in (1, 2, 3):
        ws.cell(row=2, column=col).fill = _fill(MID_BG)
    ws.row_dimensions[2].height = 20

    # Column headers row 3
    headers = [
        "#", "URL", "Channel",
        "R", "G", "B",
        "H°", "S%", "V%",
        "L*", "a*", "b*",
        "L", "C", "H°(LCH)",
    ]
    stat_names = ["mean", "median", "std", "min", "max"]
    full_headers = ["#", "URL", "Stat"] + [
        f"{ch}\n{s}"
        for ch in ["R","G","B","H°","S%","V%","L*","a*","b*","L","C","H°(LCH)"]
        for s in stat_names
    ]
    # Simpler: one row = one stat per space
    stat_headers = ["#", "URL", "Filename", "Space", "Channel",
                    "Mean", "Median", "Min", "Max", "Std"]
    for ci, h in enumerate(stat_headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font      = _header_font(size=9)
        c.fill      = _fill(MID_BG)
        c.alignment = _center(wrap=True)
        c.border    = _border()
    ws.row_dimensions[3].height = 22

    row = 4
    SPACE_COLORS = {
        "RGB": "D5E8D4", "HSV": "DAE8FC",
        "Lab": "FFF2CC", "LCH": "F8CECC",
    }
    for img_i, res in enumerate(results, 1):
        if res["status"] != "ok" or not res["stats"]:
            continue
        for space, channels in res["stats"].items():
            sp_bg = SPACE_COLORS.get(space, ROW_ODD)
            for ch, vals in channels.items():
                def wc(col, val, bold=False, bg=sp_bg):
                    c = ws.cell(row=row, column=col, value=val)
                    c.font      = Font(name="Arial", size=9, bold=bold,
                                       color="2C3E50")
                    c.fill      = _fill(bg)
                    c.alignment = _center()
                    c.border    = _border()

                wc(1,  img_i, bold=True)
                c2 = ws.cell(row=row, column=2, value=res["url"])
                c2.hyperlink  = res["url"]
                c2.font       = Font(name="Arial", size=9, color=LINK_FONT,
                                     underline="single")
                c2.fill       = _fill(sp_bg)
                c2.alignment  = _left(wrap=True)
                c2.border     = _border()

                c3 = ws.cell(row=row, column=3, value=res["filename"])
                c3.font       = Font(name="Arial", size=9, color="2C3E50")
                c3.fill       = _fill(sp_bg)
                c3.alignment  = _left()
                c3.border     = _border()

                wc(4,  space, bold=True)
                wc(5,  ch,    bold=True)
                wc(6,  vals["mean"])
                wc(7,  vals["median"])
                wc(8,  vals["min"])
                wc(9,  vals["max"])
                wc(10, vals["std"])
                ws.row_dimensions[row].height = 16
                row += 1

    _set_col_widths(ws, {
        "A": 5,  "B": 40, "C": 22,
        "D": 8,  "E": 8,
        "F": 10, "G": 10, "H": 10, "I": 10, "J": 10,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Main pipeline
# ─────────────────────────────────────────────────────────────────────────────

def run_batch(
    url_source: str,
    k: int = 6,
    output_dir: Path = Path("."),
    max_pixels: int = 300_000,
    save_png: bool = False,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    urls = load_url_list(url_source)

    results = []

    for i, url in enumerate(urls, 1):
        filename = url.split("/")[-1].split("?")[0] or f"image_{i}"
        print(f"\n[{i}/{len(urls)}] {filename}")
        print(f"    {url}")

        entry = {
            "url":      url,
            "filename": filename,
            "width":    0,
            "height":   0,
            "status":   "error",
            "stats":    None,
            "clusters": [],
            "error":    None,
        }

        try:
            img = download_image(url)
            entry["width"], entry["height"] = img.size

            # Optionally save PNG visualisation
            if save_png:
                tmp_png = output_dir / filename
                img.save(tmp_png)
                stats, clusters = analyze(
                    tmp_png, k=k,
                    output_dir=output_dir,
                    max_pixels=max_pixels,
                    save_vis=True,
                    save_json=False,
                )
            else:
                # Analyse in-memory (no disk write)
                import math
                img_small = img
                total = img.size[0] * img.size[1]
                if total > max_pixels:
                    scale = math.sqrt(max_pixels / total)
                    nw = max(1, int(img.size[0] * scale))
                    nh = max(1, int(img.size[1] * scale))
                    img_small = img.resize((nw, nh), Image.LANCZOS)

                pixels = np.array(img_small).reshape(-1, 3)
                stats    = image_statistics(pixels)
                clusters, _ = cluster_pixels(pixels, k=k, use_lab=True)

            entry["stats"]    = stats
            entry["clusters"] = clusters
            entry["status"]   = "ok"
            print(f"    ✓  {len(clusters)} clusters extracted")

        except Exception as exc:
            entry["error"] = str(exc)
            print(f"    ✗  {exc}")

        results.append(entry)
        time.sleep(0.2)   # be polite to servers

    # ── Export ───────────────────────────────────────────────────────────────
    excel_path = output_dir / "color_summary_batch.xlsx"
    print(f"\n[→] Building Excel report…")
    build_excel(results, excel_path)

    ok_count = sum(1 for r in results if r["status"] == "ok")
    print(f"\n    Processed : {len(results)} image(s)")
    print(f"    Success   : {ok_count}")
    print(f"    Failed    : {len(results) - ok_count}")
    return excel_path


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def build_parser():
    p = argparse.ArgumentParser(
        prog="batch_from_urls",
        description=(
            "Read image URLs from a .txt file (local or GitHub),\n"
            "analyse colors, and export results to Excel."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
examples:
  python batch_from_urls.py urls.txt
  python batch_from_urls.py urls.txt -k 8 -o ./results
  python batch_from_urls.py https://raw.githubusercontent.com/YOU/REPO/main/urls.txt
  python batch_from_urls.py urls.txt --save-png      # also save PNG per image
        """,
    )
    p.add_argument("source", metavar="SOURCE",
                   help="Local .txt file or GitHub URL pointing to a URL list")
    p.add_argument("-k", "--clusters", type=int, default=6,
                   metavar="N", help="Number of color clusters (default: 6)")
    p.add_argument("-o", "--output", default=".",
                   metavar="DIR", help="Output directory (default: current)")
    p.add_argument("--max-pixels", type=int, default=300_000,
                   metavar="N", help="Max pixels for analysis (default: 300000)")
    p.add_argument("--save-png", action="store_true",
                   help="Also save per-image PNG visualisations")
    return p


def main():
    args = build_parser().parse_args()
    run_batch(
        url_source=args.source,
        k=args.clusters,
        output_dir=Path(args.output),
        max_pixels=args.max_pixels,
        save_png=args.save_png,
    )


if __name__ == "__main__":
    main()
